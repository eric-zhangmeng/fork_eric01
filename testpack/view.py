from django.http import HttpResponse
from django.shortcuts import render,HttpResponse,redirect,HttpResponseRedirect
from django.contrib import messages
import xlrd, xlwt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from testpackdb.models import Cases
import datetime, time, json, os, sys

listState=['active','inactive','being fixed','awaiting test','new']
listModule=['mx-10g','fx-10g','dx-10g','mx2-10g','fx2-10g','dx2-10g','mx3-10g','fx3-40g','dx3-100g','mx4-10g','mx4-25g','mx4-40g']

def suite(request):
    context = {}
    return render(request, './View/suite.html', context)

def table1(request):
    context = {}
    context['test1'] = 'Project Name'
    context['test2'] = 'SD-WAN-BDC'
    return render(request, './View/table1.html', context)

def testcase(request):
    context = {}
    caseResult = Cases.objects.all().order_by("modifyTime")
    context['caseInfoList'] = []
    caseNumber = len(caseResult)
    caseNumberAll = len(caseResult)
    casePage = int((caseNumber-1)/10 + 1)
    for i in range(1,(caseNumber+1)):
        index = caseNumber-i
        context['caseInfoList'].append((caseResult[index].id,caseResult[index].name,caseResult[index].miscFile,caseResult[index].module,caseResult[index].priority))
    return render(request, './View/testcase.html', context)

def testbed(request):
    context = {}
    return render(request, './View/testbed.html', context)

def testpackCaseContent(request):
    context = {}
    return render(request, './View/testpackCaseContent.html', context)

def stc_testbed(request):
    context = {}
    return render(request, './View/stc_testbed.html', context)

def result_testpack(request):
    context = {}
    return render(request, './View/result_testpack.html', context)
 
def main(request):
    context = {}
    return render(request, './View/main.html', context)

 
def caseFirst(request):
    context = {}
    caseResult = Cases.objects.all().order_by("modifyTime")
    context['caseInfoList'] = []
    caseNumber = len(caseResult)
    caseNumberAll = len(caseResult)
    for i in range(1,(caseNumber+1)):
        index = caseNumber-i
        context['caseInfoList'].append((caseResult[index].id,caseResult[index].name,caseResult[index].modifyTime,caseResult[index].state))
    return render(request, './View/caseFirst.html', context)


def casePage(request):
    pageNumber = request.GET.get('pageNumber',123456)
    keyWord = request.GET.get('keyWord','noKeyWord')
    Filter_modifyTime = request.GET.get('Filter_modifyTime','noModifyTime')
    Filter_module = request.GET.get('Filter_module','noModule')
    Filter_state = request.GET.get('Filter_state','noState')
    context = {}
    context['keyWord'] = keyWord
    sessionInfo = request.session.get('myuser')
    username = sessionInfo['username']
    userRole = Users.objects.get(userName=username).role
    context['userRole'] = userRole
    context['currentUser'] = username
    context['Filter_modifyTime'] = Filter_modifyTime
    context['Filter_module'] = Filter_module
    context['Filter_state'] = Filter_state
    now = datetime.datetime.now()
    caseResult = Cases.objects.all().order_by("modifyTime")
    # return HttpResponse(keyWord)
    caseNumberAll = len(caseResult)
    if keyWord != 'noKeyWord':
        keyWordList = keyWord.split(' ')
        caseResult = Cases.objects.filter(name__contains=keyWordList[0]) 
        for i in range(1,len(keyWordList)):
            caseResult = caseResult.filter(name__contains=keyWordList[i]) 
    if Filter_modifyTime == 'Today':
        # return HttpResponse('get filter')
        today = now + datetime.timedelta(days=-1)
        caseResult = caseResult.filter(modifyTime__gte=today) 
    elif Filter_modifyTime == 'This Week':
        lastWeek = now + datetime.timedelta(days=-7)
        caseResult = caseResult.filter(modifyTime__gte=lastWeek) 
    elif Filter_modifyTime == 'This Month':
        lastMonth= now + datetime.timedelta(days=-30)
        caseResult = caseResult.filter(modifyTime__gte=lastMonth) 
    if Filter_module != 'noModule' and Filter_module != 'All' and Filter_module != '':
        caseResult = caseResult.filter(module__contains=Filter_module) 
    if Filter_state != 'noState' and Filter_state != 'All' and Filter_state != '':
        caseResult = caseResult.filter(state__contains=Filter_state)
    context['caseInfoList'] = []
    context['pageInfoList'] = []
    caseNumber = len(caseResult)
    caseCurrentNumber = caseNumber-(int(pageNumber)-1)*10
    casePage = int((caseNumber-1)/10 + 1)
    context['caseTotalNumber'] = caseNumber
    context['caseNumberAll'] = caseNumberAll
    context['caseTotalPage'] = casePage
    context['caseTotalPageList'] = range(2,(casePage+1))
    context['currentPage'] = int(pageNumber)
    for i in range(1,11):
        index = caseCurrentNumber-i
        if index >= 0:
            context['caseInfoList'].append((caseResult[index].id,caseResult[index].name,caseResult[index].modifyTime,caseResult[index].state))
    if casePage > 10:
        if int(pageNumber) <= 6:
            for i in range(1,(int(pageNumber)+3)):
                context['pageInfoList'].append(i)
            context['pageInfoList'].append('......')
            context['pageInfoList'].append(casePage-2)
            context['pageInfoList'].append(casePage-1)
            context['pageInfoList'].append(casePage)
        elif int(pageNumber) >= (casePage - 5):
            for i in range(1,4):
                context['pageInfoList'].append(i)
            context['pageInfoList'].append('......')
            for i in range((int(pageNumber)-2),(casePage + 1)):
                context['pageInfoList'].append(i)
        else:
            for i in range(1,4):
                context['pageInfoList'].append(i)
            context['pageInfoList'].append('......')
            for i in range((int(pageNumber)-2),(int(pageNumber)+3)):
                context['pageInfoList'].append(i)
            context['pageInfoList'].append('......')
            context['pageInfoList'].append(casePage-2)
            context['pageInfoList'].append(casePage-1)
            context['pageInfoList'].append(casePage)
    else:
        for i in range(1,(casePage+1)):
                context['pageInfoList'].append(i)
    # return HttpResponse(keyWord)
    return render(request, './View/casePage.html', context)


def caseContent(request):
    if request.method == "POST":
        currentId = request.POST.get('id','noId')
        if currentId == 'noId':
            caseName = request.POST.get('caseName','noCaseName')
            priority = request.POST.get('priority','noPriority')
            state = request.POST.get('state','noState')
            team = request.POST.get('team','noTeam')
            miscFile = request.POST.get('miscFile','noMiscFile')
            moduleList = request.POST.getlist('module',['noModule'])
            module = ','.join(moduleList)
            variable = request.POST.get('variable','noVariable')
            if (len(Cases.objects.filter(name=caseName))) == 0:
                test1 = Cases(name=caseName, priority=priority, state=state, team=team, miscFile=miscFile, module=module, variable=variable, modifyTime=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                test1.save()
            else:
                Cases.objects.filter(name=caseName).update(name=caseName, priority=priority, state=state, team=team, miscFile=miscFile, module=module, variable=variable, modifyTime=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            return HttpResponseRedirect('/caseFirst?saveOk=yes')
        else:
            caseName = request.POST.get('caseName','noCaseName')
            priority = request.POST.get('priority','noPriority')
            state = request.POST.get('state','noState')
            team = request.POST.get('team','noTeam')
            miscFile = request.POST.get('miscFile','noMiscFile')
            moduleList = request.POST.getlist('module',['noModule'])
            module = ','.join(moduleList)
            variable = request.POST.get('variable','noVariable')
            Cases.objects.filter(id=currentId).update(name=caseName, priority=priority, state=state, team=team, miscFile=miscFile, module=module, variable=variable, modifyTime=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            return HttpResponseRedirect('/caseFirst?saveOk=yes')
    caseId = request.GET.get('caseId','noId')
    context = {}
    sessionInfo = request.session.get('myuser')
    username = sessionInfo['username']
    context['currentUser'] = username
    if caseId == 'noId':
        context['caseName'] = ''
        context['priority'] = ''
        context['state'] = ''
        context['team'] = ''
        context['miscFile'] = ''
        context['module'] = ''
        context['variable'] = ''
        context['addCase'] = 'yes'
    else:
        caseResult = Cases.objects.get(id=caseId)
        context['id'] = caseResult.id
        context['caseName'] = caseResult.name
        context['priority'] = caseResult.priority
        context['state'] = caseResult.state
        context['team'] = caseResult.team
        context['variableList'] = []
        numVariable = 0
        infoVariable = caseResult.variable
        if infoVariable != None and infoVariable != '' and infoVariable != 'noVariable':
            infoVariableList = infoVariable.split(',')
            for i in range(len(infoVariableList)):
                listNameAndValue = infoVariableList[i].split(':')
                numVariable+=1
                j = i+1
                idForVariableType= 'VariableTypeDelete_00' + str(j)
                idForVariableValue= 'VariableValueDelete_00' + str(j)
                idForDeleteVariable= 'Delete_00' + str(j)
                idForVariableLabel= 'VariableLabelDelete_00' + str(j)
                idForVariableBr = 'VariableBrDelete_00' + str(j)
                if len(infoVariableList) == 2:
                    context['variableList'].append((listNameAndValue[0],listNameAndValue[1],idForVariableType,idForVariableValue,idForDeleteVariable,idForVariableLabel,idForVariableBr))
        context['miscFileList'] = []
        numMiscFile = 0
        infomiscFile = caseResult.miscFile
        if infomiscFile != None and infomiscFile != '' and infomiscFile != 'noMiscFile':
            infomiscFileList = infomiscFile.split(',')
            for i in range(len(infomiscFileList)):
                numMiscFile+=1
                j = i+1
                idMiscFile= 'MiscFile_00' + str(j)
                idDeleteButton= 'File_00' + str(j)
                idLabel= 'MiscLabelFile_00' + str(j)
                idBr = 'MiscBrFile_00' + str(j)
                context['miscFileList'].append((infomiscFileList[i],idMiscFile,idDeleteButton,idLabel,idBr))
        context['listChosenModule'] = []
        context['listAllModule'] = []
        infoModule = caseResult.module
        if infoModule != None and infoModule != '' and infoModule != 'noModule':
            infoModuleList = infoModule.split(',')
            for i in range(len(listModule)):
                if infoModuleList.count(listModule[i]) != 0:
                    context['listChosenModule'].append('true')
                    context['listAllModule'].append('none')
                else:
                    context['listChosenModule'].append('none')
                    context['listAllModule'].append('true')
        else:
            for i in range(len(listModule)):
                context['listChosenModule'].append('none')
                context['listAllModule'].append('true')
        context['variable'] = caseResult.variable
        context['modifyTime'] = caseResult.modifyTime
        context['viewCase'] = 'yes'
    return render(request, './View/caseContent.html', context)

def deleteCases(request):
    cases = Cases.objects.all()
    for case in cases:
        test1.delete()
    return HttpResponseRedirect('/caseFirst?deleteOk=yes')


def importCases(request):
    context = {}
    if request.method == "POST":
        fileHandle = request.FILES.get('fileUpload')
        workBook = xlrd.open_workbook(filename=None, file_contents=fileHandle.read())
        table = workBook.sheets()[0]
        rowNumber = table.nrows
        colNumber = table.ncols
        if table.row_values(0)[0] == 'name' and table.row_values(0)[1] == 'priority' and table.row_values(0)[2] == 'state' and table.row_values(0)[3] == 'team':
            # return HttpResponse(rowNumber)
            for row in range(1,rowNumber):
                if listState.count(table.row_values(row)[2]) != 0:
                    if table.row_values(row)[6]!='':
                        strVariable = table.row_values(row)[6].replace('(','')
                        strVariable = table.row_values(row)[6].replace(')','')
                    else:
                        strVariable = ''
                    if (len(Cases.objects.filter(name=table.row_values(row)[0]))) == 0:
                        caseTemp = Cases(name=table.row_values(row)[0],priority=table.row_values(row)[1],state=table.row_values(row)[2],team=table.row_values(row)[3], miscFile=table.row_values(row)[4],module=table.row_values(row)[5],variable=strVariable,modifyTime=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                        caseTemp.save()
                        # return HttpResponse('correct')
                        context['errorImportSuccess'] = 'yes'
                    else:
                        Cases.objects.filter(name=table.row_values(row)[0]).update(name=table.row_values(row)[0],priority=table.row_values(row)[1],state=table.row_values(row)[2],team=table.row_values(row)[3],miscFile=table.row_values(row)[4],module=table.row_values(row)[5],variable=strVariable,modifyTime=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                        context['errorImportSuccess'] = 'yes'
                else:
                    context['errorImportState'] = 'yes'
                    return render(request, './View/error.html', context)
            return render(request, './View/error.html', context)
        else:
            context['errorImportColumn'] = 'yes'
            return render(request, './View/error.html', context)
    return render(request, './View/importCases.html', context)

def actionOfCase(request):
    caseAction = request.POST.get('caseAction','noAction')
    casesForAction= request.POST.get('casesForAction','noCaseStr')
    numCaseSelect= request.POST.get('numCaseSelect','noNumCaseSelect')
    context = {}
    context['casesForAction'] = casesForAction
    now = datetime.datetime.now()
    if casesForAction != 'noCaseStr' and casesForAction != '':
        if numCaseSelect != '0':
            caseResult = Cases.objects.all().order_by("modifyTime")
            if keyWord != 'noKeyWord':
                keyWordList = keyWord.split(' ')
                caseResult = Cases.objects.filter(name__contains=keyWordList[0]) 
                for i in range(1,len(keyWordList)):
                    caseResult = caseResult.filter(name__contains=keyWordList[i]) 
            if Filter_modifyTime == 'Today':
                today = now + datetime.timedelta(days=-1)
                caseResult = caseResult.filter(modifyTime__gte=today) 
            elif Filter_modifyTime == 'This Week':
                lastWeek = now + datetime.timedelta(days=-7)
                caseResult = caseResult.filter(modifyTime__gte=lastWeek) 
            elif Filter_modifyTime == 'This Month':
                lastMonth= now + datetime.timedelta(days=-30)
                caseResult = caseResult.filter(modifyTime__gte=lastMonth) 
            if Filter_module != 'noModule' and Filter_module != 'All' and Filter_module != '':
                caseResult = caseResult.filter(module__contains=Filter_module) 
            if Filter_state != 'noState' and Filter_state != 'All' and Filter_state != '':
                caseResult = caseResult.filter(state__contains=Filter_state)
            if int(numCaseSelect)<=len(caseResult):
                tempCaseNum = int(numCaseSelect)
            else:
                tempCaseNum = len(caseResult)
            caselist = []
            for i in range(0,tempCaseNum):
                caselist.append(caseResult[i].name)
        else:
            caselist = casesForAction.split(',')
    else:
        return HttpResponse('need to select some cases firstly, please go back to last page now!')
    if caseAction == 'noAction' or caseAction == '--------':
        return HttpResponse('no action at all, please go back to last page')
    elif caseAction == 'Edit':
        return render(request, './View/caseEdit.html', context)
    elif caseAction == 'Export':
        if os.path.exists("./static/file/exportSelectedCases.xlsx"):
            os.remove("./static/file/exportSelectedCases.xlsx")
        workBook = Workbook()
        workSheet = workBook.active
        workSheet.cell(row=1, column=1).value = 'name'
        workSheet.cell(row=1, column=2).value = 'priority'
        workSheet.cell(row=1, column=3).value = 'state'
        workSheet.cell(row=1, column=4).value = 'team'
        workSheet.cell(row=1, column=5).value = 'miscFile'
        workSheet.cell(row=1, column=6).value = 'module'
        workSheet.cell(row=1, column=7).value = 'variable'
        # return HttpResponse('mark1')
        for i in range(0,len(caselist)):
            rowTemp = i+2
            caseResult = Cases.objects.filter(name=caselist[i]) 
            variableList = caseResult[0].variable.split(',')
            for j in range(len(variableList)):
                variableList[j] = '(' + variableList[j] + ')'            
            newVarStr = ','.join(variableList)
            workSheet.cell(row=rowTemp, column=1).value = caseResult[0].name
            workSheet.cell(row=rowTemp, column=2).value = caseResult[0].priority
            workSheet.cell(row=rowTemp, column=3).value = caseResult[0].state
            workSheet.cell(row=rowTemp, column=4).value = caseResult[0].team
            workSheet.cell(row=rowTemp, column=5).value = caseResult[0].miscFile
            workSheet.cell(row=rowTemp, column=6).value = caseResult[0].module
            workSheet.cell(row=rowTemp, column=7).value = newVarStr
        workBook.save(filename="./static/file/exportSelectedCases.xlsx")
        return HttpResponseRedirect('/static/file/exportSelectedCases.xlsx')
    elif caseAction == 'Duplicate':
        for i in range(0,len(caselist)):
            caseName = Cases.objects.get(name=caselist[i]).name + '_Duplicate'
            priority = Cases.objects.get(name=caselist[i]).priority
            state = Cases.objects.get(name=caselist[i]).state
            team = Cases.objects.get(name=caselist[i]).team
            miscFile = Cases.objects.get(name=caselist[i]).miscFile
            module = Cases.objects.get(name=caselist[i]).module
            variable = Cases.objects.get(name=caselist[i]).variable
            tempCase = Cases(name=caseName, priority=priority, state=state, team=team, miscFile=miscFile, module=module, variable=variable, modifyTime=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            tempCase.save()
        context = {}
        sessionInfo = request.session.get('myuser')
        caseResult = Cases.objects.all().order_by("modifyTime")
        context['caseInfoList'] = []
        caseNumber = len(caseResult)
        casePage = int((caseNumber-1)/10 + 1)
        context['caseTotalNumber'] = caseNumber
        context['caseTotalPage'] = casePage
        context['caseTotalPageList'] = range(2,(casePage+1))
        if casePage == 1 and caseNumber != 0:
            for i in range(1,(caseNumber+1)):
                index = caseNumber-i
                context['caseInfoList'].append((caseResult[index].id,caseResult[index].name,caseResult[index].modifyTime,caseResult[index].state))
        elif caseNumber != 0:
            for i in range(1,11):
                index = caseNumber-i
                context['caseInfoList'].append((caseResult[index].id,caseResult[index].name,caseResult[index].modifyTime,caseResult[index].state))
        return render(request, './View/caseFirst.html', context)
    elif caseAction == 'Delete':
        for i in range(0,len(caselist)):
            Cases.objects.get(name=caselist[i]).delete()
        context = {}
        caseResult = Cases.objects.all().order_by("modifyTime")
        context['caseInfoList'] = []
        caseNumber = len(caseResult)
        caseNumberAll = len(caseResult)
        casePage = int((caseNumber-1)/10 + 1)
        context['caseTotalNumber'] = caseNumber
        context['caseNumberAll'] = caseNumberAll
        context['caseTotalPage'] = casePage
        context['caseTotalPageList'] = range(2,(casePage+1))
        if casePage == 1 and caseNumber != 0:
            for i in range(1,(caseNumber+1)):
                index = caseNumber-i
                context['caseInfoList'].append((caseResult[index].id,caseResult[index].name,caseResult[index].modifyTime,caseResult[index].state))
        elif caseNumber != 0:
            for i in range(1,11):
                index = caseNumber-i
                context['caseInfoList'].append((caseResult[index].id,caseResult[index].name,caseResult[index].modifyTime,caseResult[index].state))
        return render(request, './View/caseFirst.html', context)